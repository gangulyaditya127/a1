import os
import shutil
import requests
from datetime import datetime
from openpyxl import load_workbook

# =====================
# Config
# =====================
FOLDER_PATH = os.path.join(os.path.dirname(__file__), "..", "data", "input_excels")
FOLDER_PATH = os.path.abspath(FOLDER_PATH)

JAVA_SERVICE_URL = "https://localhost:8082/ismartService/user/parse"
LOG_FOLDER = os.path.join(os.path.dirname(__file__), "..", "logs")
LOG_FILE = os.path.join(LOG_FOLDER, "summary.log")
PROCESSED_FOLDER = os.path.join(os.path.dirname(__file__), "..", "processed")

REQUIRED_SHEETS = {"Engagement Overview", "Hotspots", "ITSM Practices"}

# Ensure folders exist
os.makedirs(LOG_FOLDER, exist_ok=True)
os.makedirs(PROCESSED_FOLDER, exist_ok=True)

# =====================
# Utility Functions
# =====================
def write_log(message):
    print(message)
    with open(LOG_FILE, "a", encoding="utf-8") as f:
        f.write(message + "\n")


def move_file(file_path, category):
    """Move file to category subfolder under processed/"""
    category_folder = os.path.join(PROCESSED_FOLDER, category)
    os.makedirs(category_folder, exist_ok=True)
    dest_path = os.path.join(category_folder, os.path.basename(file_path))
    shutil.move(file_path, dest_path)
    write_log(f"📂 Moved {os.path.basename(file_path)} → {category}/")


def validate_excel(file_path):
    try:
        workbook = load_workbook(file_path, read_only=True)
        sheet_names = set(workbook.sheetnames)
        workbook.close()

        missing = REQUIRED_SHEETS - sheet_names
        if missing:
            return False, f"Missing sheets: {', '.join(missing)}"

        return True, "All mandatory sheets present"

    except Exception as e:
        return False, f"Excel validation error: {e}"


# =====================
# Processing Helpers
# =====================
def handle_validation_failure(file_name, file_path, message, validation_failures):
    validation_failures[file_name] = message
    write_log(f"❌ Skipping {file_name} → {message}")
    move_file(file_path, "validation_failures")


def call_java_service(file_name, file_path, success_files, service_failures):
    write_log(f"Processing: {file_name}")

    with open(file_path, "rb") as f:
        response = requests.post(
            JAVA_SERVICE_URL,
            files={"file": f},
            verify=False
        )

    if response.status_code == 200:
        success_files.append(file_name)
        write_log(f"✅ Success: {file_name}")
        move_file(file_path, "success")
    else:
        service_failures[file_name] = f"HTTP {response.status_code} - {response.text}"
        write_log(f"❌ Service failed: {file_name} (status {response.status_code})")
        move_file(file_path, "service_failures")


def process_single_file(
    file_name,
    success_files,
    validation_failures,
    service_failures,
    other_errors
):
    file_path = os.path.join(FOLDER_PATH, file_name)

    # Step 1: Validation
    is_valid, validation_msg = validate_excel(file_path)
    if not is_valid:
        handle_validation_failure(
            file_name,
            file_path,
            validation_msg,
            validation_failures
        )
        return

    write_log(f"✅ {file_name} → {validation_msg}")

    # Step 2: Service call
    try:
        call_java_service(
            file_name,
            file_path,
            success_files,
            service_failures
        )
    except Exception as e:
        other_errors[file_name] = str(e)
        write_log(f"⚠️ Error processing {file_name}: {e}")
        move_file(file_path, "other_errors")


def write_summary(success_files, validation_failures, service_failures, other_errors):
    write_log("\n==== Summary ====")

    if success_files:
        write_log("✔️ Completed successfully:")
        for f in success_files:
            write_log(f"   - {f}")
    else:
        write_log("No files completed successfully.")

    if validation_failures:
        write_log("\n❌ Validation Failures:")
        for f, reason in validation_failures.items():
            write_log(f"   - {f} : {reason}")

    if service_failures:
        write_log("\n❌ Service Failures:")
        for f, reason in service_failures.items():
            write_log(f"   - {f} : {reason}")

    if other_errors:
        write_log("\n⚠️ Other Errors:")
        for f, reason in other_errors.items():
            write_log(f"   - {f} : {reason}")

    write_log("===== Run finished =====\n")


# =====================
# Main Orchestrator
# =====================
def process_excel_files():
    success_files = []
    validation_failures = {}
    service_failures = {}
    other_errors = {}

    write_log(
        f"\n===== Run started at {datetime.now().strftime('%Y-%m-%d %H:%M:%S')} ====="
    )

    for file_name in os.listdir(FOLDER_PATH):
        if file_name.endswith(".xlsx"):
            process_single_file(
                file_name,
                success_files,
                validation_failures,
                service_failures,
                other_errors
            )

    write_summary(
        success_files,
        validation_failures,
        service_failures,
        other_errors
    )

if __name__ == "__main__":
    process_excel_files()